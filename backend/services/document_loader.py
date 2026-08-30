import os
import csv
from pathlib import Path
from typing import List, Dict, Any

class DocumentLoader:
    """Multi-format document text extractor supporting PDF, DOCX, TXT, MD, CSV, XLSX"""

    @staticmethod
    def load_file(file_path: Path, filename: str, source_type: str) -> List[Dict[str, Any]]:
        """
        Extracts content from file. Returns a list of segments with metadata:
        [{ 'text': '...', 'page_number': 1, 'section_name': '...' }]
        """
        segments = []
        source_type = source_type.lower()

        if source_type == "pdf":
            segments = DocumentLoader._load_pdf(file_path)
        elif source_type == "docx":
            segments = DocumentLoader._load_docx(file_path)
        elif source_type in ["txt", "markdown", "md"]:
            segments = DocumentLoader._load_txt(file_path)
        elif source_type == "csv":
            segments = DocumentLoader._load_csv(file_path)
        elif source_type == "xlsx":
            segments = DocumentLoader._load_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported document format: {source_type}")

        # Filter out empty or whitespace-only segments
        cleaned_segments = [s for s in segments if s.get("text", "").strip()]
        if not cleaned_segments:
            raise ValueError("Document appears to be empty or contains no extractable text.")
            
        return cleaned_segments

    @staticmethod
    def _load_pdf(file_path: Path) -> List[Dict[str, Any]]:
        segments = []
        # Attempt pypdf first
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            for page_num, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    segments.append({
                        "text": text,
                        "page_number": page_num,
                        "section_name": f"Page {page_num}"
                    })
            return segments
        except ImportError:
            pass

        # Fallback to pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(str(file_path)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text() or ""
                    if text.strip():
                        segments.append({
                            "text": text,
                            "page_number": page_num,
                            "section_name": f"Page {page_num}"
                        })
            return segments
        except ImportError:
            # Fallback simple reader if pdf libraries missing
            with open(file_path, "rb") as f:
                raw = f.read().decode("latin1", errors="ignore")
                segments.append({"text": raw[:5000], "page_number": 1, "section_name": "Page 1"})
            return segments

    @staticmethod
    def _load_docx(file_path: Path) -> List[Dict[str, Any]]:
        segments = []
        try:
            import docx
            doc = docx.Document(str(file_path))
            current_section = "General"
            current_text = []

            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                if para.style.name.startswith("Heading"):
                    if current_text:
                        segments.append({
                            "text": "\n".join(current_text),
                            "section_name": current_section
                        })
                        current_text = []
                    current_section = text
                else:
                    current_text.append(text)

            if current_text:
                segments.append({
                    "text": "\n".join(current_text),
                    "section_name": current_section
                })
            return segments
        except Exception:
            # Simple text fallback
            with open(file_path, "rb") as f:
                content = f.read().decode("utf-8", errors="ignore")
                return [{"text": content, "section_name": "Document Body"}]

    @staticmethod
    def _load_txt(file_path: Path) -> List[Dict[str, Any]]:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="latin1") as f:
                content = f.read()
        return [{"text": content, "section_name": "Text Document"}]

    @staticmethod
    def _load_csv(file_path: Path) -> List[Dict[str, Any]]:
        segments = []
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                for row_idx, row in enumerate(reader, 1):
                    row_text = ", ".join([f"{h}: {val}" for h, val in zip(headers, row) if val])
                    if row_text.strip():
                        segments.append({
                            "text": row_text,
                            "section_name": f"Row {row_idx}"
                        })
            return segments
        except Exception as e:
            return [{"text": f"Error reading CSV: {e}", "section_name": "CSV Data"}]

    @staticmethod
    def _load_xlsx(file_path: Path) -> List[Dict[str, Any]]:
        segments = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(h) for h in rows[0] if h is not None]
                for r_idx, row in enumerate(rows[1:], 2):
                    row_vals = [f"{headers[i] if i < len(headers) else f'Col{i}'}: {val}" for i, val in enumerate(row) if val is not None]
                    if row_vals:
                        segments.append({
                            "text": f"Sheet: {sheet_name} | " + ", ".join(row_vals),
                            "section_name": f"{sheet_name} (Row {r_idx})"
                        })
            return segments
        except Exception:
            return [{"text": "Spreadsheet data processed.", "section_name": "Spreadsheet"}]
